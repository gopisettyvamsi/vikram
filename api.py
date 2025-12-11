from flask import Flask, request, jsonify, send_file, abort
from flask_cors import CORS
from functools import wraps
import json
import csv
import io
import pandas as pd
from app import generate_test_cases

app = Flask(__name__)
CORS(app)  # Enable CORS for external clients

# Configure allowed IP addresses
ALLOWED_IPS = [
    '127.0.0.1',      # localhost
    '::1',            # localhost IPv6
    '125.21.51.10',   # Your IP address
    '192.168.0.127',  # Local network IP
]

def require_ip_whitelist(f):
    """Decorator to restrict access to whitelisted IP addresses"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        client_ip = request.remote_addr

        # Print to console for debugging
        print(f"\n{'='*60}")
        print(f"IP ACCESS CHECK - Flask API")
        print(f"{'='*60}")
        print(f"Current IP Address: {client_ip}")
        print(f"Allowed IP Addresses: {ALLOWED_IPS}")
        print(f"Access Allowed: {client_ip in ALLOWED_IPS}")
        print(f"Endpoint: {request.endpoint}")
        print(f"{'='*60}\n")

        # Check if client IP is in allowed list
        if client_ip not in ALLOWED_IPS:
            print(f"[BLOCKED] Access denied for IP: {client_ip}")
            return jsonify({
                "error": "Access denied",
                "message": f"Your IP address ({client_ip}) is not authorized to access this resource",
                "allowed_ips": ALLOWED_IPS,
                "your_ip": client_ip
            }), 403

        print(f"[ALLOWED] Access granted for IP: {client_ip}")
        return f(*args, **kwargs)
    return decorated_function

@app.route('/api/generate', methods=['POST'])
@require_ip_whitelist
def generate():
    """Generate test cases endpoint"""
    try:
        data = request.get_json()
        module = data.get('module', '')

        if not module.strip():
            return jsonify({"error": "Module description is required"}), 400

        result = generate_test_cases(module)

        if "error" in result:
            return jsonify(result), 500

        return jsonify(result), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/export/json', methods=['POST'])
@require_ip_whitelist
def export_json():
    """Export test cases as JSON"""
    try:
        data = request.get_json()
        test_cases = data.get('test_cases', {})

        json_str = json.dumps(test_cases, indent=2)

        return jsonify({
            "data": json_str,
            "filename": "test_cases.json"
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/export/csv', methods=['POST'])
@require_ip_whitelist
def export_csv():
    """Export test cases as CSV with title and status"""
    try:
        data = request.get_json()
        test_cases = data.get('test_cases', {})

        # Create CSV in memory
        output = io.StringIO()
        fieldnames = ["ID", "Title", "Scenario", "Type", "Steps", "Expected Result", "Status"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)

        writer.writeheader()

        for tc in test_cases.get("test_cases", []):
            writer.writerow({
                "ID": tc.get("id", ""),
                "Title": tc.get("title", ""),
                "Scenario": tc.get("scenario", ""),
                "Type": tc.get("type", ""),
                "Steps": "; ".join(tc.get("steps", [])),
                "Expected Result": tc.get("expected_result", ""),
                "Status": tc.get("status", "Pending")
            })

        csv_data = output.getvalue()

        return jsonify({
            "data": csv_data,
            "filename": "test_cases.csv"
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/export/excel', methods=['POST'])
@require_ip_whitelist
def export_excel():
    """Export test cases as Excel with color-coded status"""
    try:
        data = request.get_json()
        test_cases = data.get('test_cases', {})

        # Create DataFrame
        df_data = []
        for tc in test_cases.get("test_cases", []):
            df_data.append({
                "ID": tc.get("id", ""),
                "Title": tc.get("title", ""),
                "Scenario": tc.get("scenario", ""),
                "Type": tc.get("type", ""),
                "Steps": "; ".join(tc.get("steps", [])),
                "Expected Result": tc.get("expected_result", ""),
                "Status": tc.get("status", "Pending")
            })

        df = pd.DataFrame(df_data)

        # Create Excel file in memory with formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Test Cases')

            # Get the worksheet
            from openpyxl.styles import PatternFill, Font
            worksheet = writer.sheets['Test Cases']

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Color code status column
            status_col = 7  # Column G (Status)
            for row in range(2, len(df_data) + 2):
                cell = worksheet.cell(row=row, column=status_col)
                if cell.value == "Passed":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                elif cell.value == "Failed":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)
                elif cell.value == "Pending":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500", bold=True)

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='test_cases.xlsx'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/health', methods=['GET'])
@require_ip_whitelist
def health():
    """Health check endpoint"""
    return jsonify({"status": "healthy"}), 200


if __name__ == '__main__':
    app.run(debug=True, port=5000)
