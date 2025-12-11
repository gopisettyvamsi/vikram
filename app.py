import os
import json
from dotenv import load_dotenv
from groq import Groq
import streamlit as st

# Load environment variables
load_dotenv()

# Try to get API key from Streamlit secrets first (for production), then fall back to .env (for local)
try:
    api_key = st.secrets["GROQ_API_KEY"]
except (FileNotFoundError, KeyError):
    api_key = os.getenv("GROQ_API_KEY")

if not api_key:
    st.error("⚠️ GROQ_API_KEY not found. Please set it in Streamlit secrets or your .env file.")
    st.stop()

client = Groq(api_key=api_key)

PROMPT_TEMPLATE = """
You are a senior QA engineer.

Generate COMPLETE test cases for the following module/page:

MODULE:
{module}

Cover:
- Functional test cases
- Negative test cases
- Boundary cases
- Validation checks
- Basic security cases

Return STRICT JSON only in this format:

{{
  "module": "{module}",
  "total_test_cases": number,
  "test_cases": [
    {{
      "id": "TC-001",
      "title": "Brief descriptive title",
      "scenario": "Detailed scenario description",
      "type": "Functional | Negative | Boundary | Security",
      "steps": ["step1", "step2"],
      "expected_result": "Expected outcome",
      "status": "Pending"
    }}
  ]
}}

Rules:
- ONLY JSON
- No explanation text
- Must be valid JSON
- title should be short (3-8 words)
- status should always be "Pending" initially
"""

def generate_test_cases(module: str):
    prompt = PROMPT_TEMPLATE.format(module=module)

    response = client.chat.completions.create(
        # model="llama3-70b-8192",
        # model="llama-3.1-70b-versatile",
        model="llama-3.1-8b-instant",


        messages=[{"role": "user", "content": prompt}],
        temperature=0.2
    )

    text = response.choices[0].message.content.strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {"error": "Invalid JSON from model", "raw": text}


def export_to_excel(test_cases: dict, filename: str = "test_cases.xlsx"):
    """Export test cases to an Excel file."""
    try:
        import pandas as pd
        from openpyxl.styles import PatternFill, Font

        # Prepare data for Excel
        data = []
        for tc in test_cases.get("test_cases", []):
            data.append({
                "ID": tc.get("id", ""),
                "Title": tc.get("title", ""),
                "Scenario": tc.get("scenario", ""),
                "Type": tc.get("type", ""),
                "Steps": "; ".join(tc.get("steps", [])),
                "Expected Result": tc.get("expected_result", ""),
                "Status": tc.get("status", "Pending")
            })

        df = pd.DataFrame(data)

        # Write to Excel with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Test Cases")

            # Get the worksheet
            worksheet = writer.sheets['Test Cases']

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Color code status column
            status_col = 7  # Column G (Status)
            for row in range(2, len(data) + 2):
                cell = worksheet.cell(row=row, column=status_col)
                if cell.value == "Passed":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                elif cell.value == "Failed":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)
                elif cell.value == "Pending":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500", bold=True)

        return {"status": "success", "filename": filename}
    except ImportError:
        return {"status": "error", "message": "pandas/openpyxl not installed. Install with: pip install pandas openpyxl"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def export_to_text(test_cases: dict, filename: str = "test_cases.txt"):
    """Export test cases to a text file."""
    try:
        with open(filename, "w", encoding="utf-8") as f:
            f.write(f"MODULE: {test_cases.get('module', 'N/A')}\n")
            f.write(f"Total Test Cases: {test_cases.get('total_test_cases', 0)}\n")
            f.write("=" * 80 + "\n\n")

            for tc in test_cases.get("test_cases", []):
                f.write(f"Test Case ID: {tc.get('id', '')}\n")
                f.write(f"Title: {tc.get('title', '')}\n")
                f.write(f"Scenario: {tc.get('scenario', '')}\n")
                f.write(f"Type: {tc.get('type', '')}\n")
                f.write(f"Steps:\n")
                for i, step in enumerate(tc.get("steps", []), 1):
                    f.write(f"  {i}. {step}\n")
                f.write(f"Expected Result: {tc.get('expected_result', '')}\n")
                f.write(f"Status: {tc.get('status', 'Pending')}\n")
                f.write("-" * 80 + "\n\n")

        return {"status": "success", "filename": filename}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def export_to_csv(test_cases: dict, filename: str = "test_cases.csv"):
    """Export test cases to a CSV file with title and status."""
    try:
        import csv

        with open(filename, "w", newline="", encoding="utf-8") as f:
            fieldnames = ["ID", "Title", "Scenario", "Type", "Steps", "Expected Result", "Status"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)

            writer.writeheader()

            for tc in test_cases.get("test_cases", []):
                writer.writerow({
                    "ID": tc.get("id", ""),
                    "Title": tc.get("title", ""),
                    "Scenario": tc.get("scenario", ""),
                    "Type": tc.get("type", ""),
                    "Steps": "; ".join(tc.get("steps", [])),
                    "Expected Result": tc.get("expected_result", ""),
                    "Status": tc.get("status", "Pending")
                })

        return {"status": "success", "filename": filename}
    except Exception as e:
        return {"status": "error", "message": str(e)}
